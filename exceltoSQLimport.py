import mysql.connector
from openpyxl import Workbook, load_workbook

wb = load_workbook('SampleExcel.xlsx')
ws = wb.active

db = mysql.connector.connect(user='root', password='password',
                              host='localhost', database='test1')

start = 2
numRows = 7

ws.insert_cols(1)

primKey = 1
for r in range(start, 9):
    ws.cell(r, 1, primKey)
    primKey += 1

wb.save('SampleExcel2.xlsx')

mycursor = db.cursor()

table = "table6"

# mycursor.execute(f"CREATE TABLE {table} (firstName VARCHAR(50), lastName VARCHAR(50), age INT, gender ENUM(\"M\",\"F\"), number INT, personID INT PRIMARY KEY)")

query = f"REPLACE INTO {table} (firstName, lastName, age, gender, number, personID) VALUES (%s, %s, %s, %s, %s, %s)"

for r in range(start, numRows):
    personID = ws.cell(r, 1).value
    firstName = ws.cell(r, 2).value
    lastName = ws.cell(r, 3).value
    age = ws.cell(r, 4).value
    gender = ws.cell(r, 5).value
    number = ws.cell(r, 6).value

    values = (firstName, lastName, age, gender, number, personID)
    mycursor.execute(query, values)

db.commit()

mycursor.execute(f"SELECT * FROM {table}")
# for x in mycursor:
#     print(x)




