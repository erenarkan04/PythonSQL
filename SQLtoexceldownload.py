import mysql.connector
from openpyxl import Workbook, load_workbook
from exceltoSQLimport import table

wb = Workbook()
ws = wb.active

db = mysql.connector.connect(user='root', password='password',
                              host='localhost', database='test1')

cursor = db.cursor()

cursor.execute(f"SELECT * FROM {table} WHERE gender = \"M\"")

# columns = cursor.execute(f"SHOW COLUMNS FROM {table}")
print(cursor.description)

# ws.cell(1, 1, )
# for x in cursor:
#     print(x)

print("hello")

